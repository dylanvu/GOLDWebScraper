# This is the main script to extract courses from GOLD, UCSB's course registration site.
# It reads a "template" Excel sheet and outputs an Excel sheet with all of the courses

import os

import openpyxl
from dotenv import load_dotenv
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from openpyxl.styles import Font

from functions import *

# Set the template file path
path = "Input Sheet.xlsx"
print("Loading up Input Excel sheet")
wb_obj = openpyxl.load_workbook(path)
print("Input Excel sheet loaded. Obtaining Courses to be Scraped")
sheet_obj = wb_obj.active

department_columns = []

cell_obj = sheet_obj.cell(row=1, column=1)
col = 1

# This loop puts all of the department headers into a list and prepares it for the scraper
while cell_obj.value is not None:
    department_columns.append(cell_obj.value.upper() + " ")
    col += 1    # Move to the next column
    cell_obj = sheet_obj.cell(row=1, column=col)

print("Courses Obtained.")
# Set the driver to headless mode
options = Options()
options.add_argument('--headless')
options.add_argument('--disable-gpu')

print("Initializing Web Driver")
driver = webdriver.Chrome("chromedriver.exe", options=options)  # Initialize web driver in headless mode

#driver = webdriver.Chrome("chromedriver.exe")       # Comment this out if you want to run in headless mode

driver.get("https://my.sa.ucsb.edu/gold/")  # Get to the GOLD website
print("On Gold")

# Avoid storing password directly into the code by creating and modifying a .env file that has the username and password stored in it.
# Import the username and password from the .env file
load_dotenv()
gold_username = os.environ.get("gold_username")
gold_password = os.environ.get("gold_password")

# Locate the GOLD username and password boxes, input the information in, and log in
username_input = driver.find_element_by_id("pageContent_userNameText")
password_input = driver.find_element_by_id("pageContent_passwordText")
username_input.send_keys(gold_username)
password_input.send_keys(gold_password)

# Overwrite the username and password
gold_username = "username"
gold_password = "password"

driver.find_element_by_id("pageContent_loginButton").click()
print("Successfully logged in")

# Navigate to the course webpage by locating the find course button and clicking on it
driver.find_element_by_id("Li1").click()

# Open the quarter dropdown menu
driver.find_element_by_name("ctl00$pageContent$quarterDropDown").click()

# Locate and click on the first quarter of the dropdown menu.
# Because of the dropdown menu being Javascript, you can't locate elements by id or name so
# you have to use xpath to navigate to the most recent quarter (and whenever you're interacting with the dropdown)
driver.find_element_by_xpath("/html/body/div[@id='body-container']"
                             "/form[@id='MainForm']"
                             "/section[@id='content']"
                             "/div[@class='row']"
                             "/main[@class='col-sm-12']"
                             "/div[@id='content']"
                             "/div[@class='formcontainer']"
                             "/div[@class='row'][2]"
                             "/div[@class='col-sm-2']"
                             "/select[@id='pageContent_quarterDropDown']/option[1]").click()

current_cell_obj = sheet_obj.cell(row=2, column=1)
department_list = []
row = 2
col = 1

print("Current Quarter Selected. Commencing Scraping...")

# For each department requested in the Excel sheet
for department in department_columns:
    # Find the xpath to the department
    department_xpath = "/html/body/div[@id='body-container']" \
                  "/form[@id='MainForm']" \
                  "/section[@id='content']" \
                  "/div[@class='row']" \
                  "/main[@class='col-sm-12']" \
                  "/div[@id='content']" \
                  "/div[@class='formcontainer']" \
                  "/div[@class='row'][2]"\
                  "/div[@class='col-sm-9']"\
                  "/div[@class='row']"\
                  "/div[@class='col-md-6 col-sm-5 find']"\
                  "/select[@id='pageContent_subjectAreaDropDown']" \
                  f"//option[@value='{department}']"

    # Locate and click on the current department on Gold, hit search, and then obtain the HTML of the page
    driver.find_element_by_xpath(department_xpath).click()
    driver.find_element_by_id("pageContent_searchButton").click()
    page_source = driver.page_source

    course_list = []

    if current_cell_obj.value.lower() == "all":   # Check to see if all courses are requested
        course_list = scrape_all_courses(page_source, department)  # Scrape every course
    else:   # Else, move onto the individual course scrape.
        soup = BeautifulSoup(page_source, 'html.parser')
        # Find all course titles in HTML format
        all_course_spans = soup.find_all("span", class_="courseTitle")
        course_titles = [unicodedata.normalize("NFKD", span.get_text()) for span in all_course_spans]  # Get the course name in text and append them (this is called list comprehension)
        while current_cell_obj.value is not None:
            # Add the course span algorithm here and turn it into an input in the function
            current_course_obj = scrape_single_course(course_titles, all_course_spans, current_cell_obj.value, department)
            if current_course_obj is not None:
                course_list.append(current_course_obj)
            row += 1    # Set the row to be the next course in the sheet
            current_cell_obj = sheet_obj.cell(row=row, column=col)  # Move onto the next course
    print("Done with " + department)
    department_list.append(Department(department, course_list))     # Add the department courses data to the list
    row = 2     # Jump back to the first class of the next department
    col += 1
    current_cell_obj = sheet_obj.cell(row=row, column=col)      # Move onto the next department
    driver.find_element_by_id("Li1").click()    # Go back to find courses. Note: can further optimize by doing a check to see if the department is the last one

print("Scraping done. Closing the Web Driver")
driver.quit()

# Setting up the output excel sheet
print("Setting up Output Excel sheet")
scraped_course_wb = openpyxl.Workbook()
ws = scraped_course_wb.active
ws.title = department_columns[0]
department_columns_copy = department_columns.copy()
department_columns_copy.pop(0)
added_ws = [scraped_course_wb.create_sheet(f"{department}") for department in department_columns_copy]

# Fill out the scraped information for each department
print("Populating Data...")
for num in range(0, len(department_columns)):
    scraped_course_wb.active = num       # Set the department worksheet to be active
    dept_ws = scraped_course_wb.active
    # Format the header
    dept_ws.cell(row=1, column=1).value = department_columns[num]
    dept_ws.cell(row=1, column=1).font = Font(bold=True)

    dept_ws.cell(row=1, column=2).value = "Available Spaces"
    dept_ws.cell(row=1, column=2).font = Font(bold=True)

    dept_ws.cell(row=1, column=3).value = "Maximum Spaces"
    dept_ws.cell(row=1, column=3).font = Font(bold=True)
    row = 2
    # For each course scraped in the department
    for unique_course in range(0, len(department_list[num].courses)):
        for lecture_num in range(0, len(department_list[num].courses[unique_course])):
            dept_ws.cell(row=row, column=1).value = concat_dept(department_list[num].courses[unique_course][lecture_num])  # Concat the name
            dept_ws.cell(row=row, column=1).font = Font(bold=True)
            dept_ws.cell(row=row, column=2).value = department_list[num].courses[unique_course][lecture_num].availability
            dept_ws.cell(row=row, column=3).value = department_list[num].courses[unique_course][lecture_num].max
            row += 1

    print(f"{department_columns[num]} is done.")

print("Script Complete!")
scraped_course_wb.save("Output Sheet.xlsx")
